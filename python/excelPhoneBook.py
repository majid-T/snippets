import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import *
from tkinter.messagebox import showinfo

win = tk.Tk()
win.title('Excel phonebook')


# Functions
def save():
    fName = fNameInput.get()
    lName = lNameInput.get()
    email = emailInput.get()
    tel = telInput.get()

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "First name"
    ws['B1'] = "Last name"
    ws['C1'] = "Email"
    ws['D1'] = "Telephone No"

    ws['A2'] = fName
    ws['B2'] = lName
    ws['C2'] = email
    ws['D2'] = tel

    wb.save(r'C:\Users\majid\Desktop\phonebook.xlsx')
    showinfo('Entry saved!', 'New contact person added!')

    clear()


def clear():
    fNameInput.delete(0, tk.END)
    lNameInput.delete(0, tk.END)
    emailInput.delete(0, tk.END)
    telInput.delete(0, tk.END)


# GUI building
fNameLabel = tk.Label(win, text='First Name: ')
fNameLabel.grid(row=0, column=0, padx=7, pady=7)

fNameInput = tk.Entry(win)
fNameInput.grid(row=0, column=1, padx=7, pady=7)
fNameInput.grid(row=0, column=1, padx=7, pady=7)

lNameLabel = tk.Label(win, text='Last Name: ')
lNameLabel.grid(row=1, column=0, padx=7, pady=7)

lNameInput = tk.Entry(win)
lNameInput.grid(row=1, column=1, padx=7, pady=7)

emailLabel = tk.Label(win, text='Email: ')
emailLabel.grid(row=2, column=0, padx=7, pady=7)

emailInput = tk.Entry(win)
emailInput.grid(row=2, column=1, padx=7, pady=7)

telLabel = tk.Label(win, text='Telephone: ')
telLabel.grid(row=3, column=0, padx=7, pady=7)

telInput = tk.Entry(win)
telInput.grid(row=3, column=1, padx=7, pady=7)


addButton = tk.Button(win, text='Add', command=save)
addButton.grid(row=4, column=0, padx=7, pady=7)

clearButton = tk.Button(win, text='Clear', command=clear)
clearButton.grid(row=4, column=1, padx=7, pady=7)


win.mainloop()
